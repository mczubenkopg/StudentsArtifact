
# -*- coding: utf-8 -*-
"""
Analiza 1. strony ankiet generowanych przez podany generator (student/teacher).
- Część A: wykrycie zaznaczonych kratek (1..6) OpenCV.
- Część B: OCR tekstu w ramkach (pytesseract, lang='pol').
- Tylko PIERWSZA strona każdego PDF (druga strona -> pytania, bez analizy).
- Export do Excel z nagłówkami A.1..A.20 i B.1..B.5 oraz podsumowaniami.

Geometria, nazwy i kolejność są dopasowane do kodu generatora.  # źródło: cop_approach_gen.py
"""

import os
import sys
import argparse
from dataclasses import dataclass
from typing import List, Tuple

import numpy as np
import cv2
from PIL import Image
from openpyxl.styles.builtins import output

try:
    from pdf2image import convert_from_path
except Exception:
    convert_from_path = None

try:
    import pytesseract
except Exception:
    pytesseract = None

from openpyxl import Workbook
from openpyxl.styles import Alignment

# ---------- Geometria zsynchronizowana z generatorem ----------
# Z generatora: A4, margins = 5 mm; fiducial_size = 5 mm; checkbox_size = 5 mm; checkbox_gap = 2 mm; col_gap = 6 mm
# oraz złożone obliczenia pozycji X_A, X_B, Y_T itd.  (patrz cop_approach_gen.py) [1](https://pgedupl-my.sharepoint.com/personal/micczube_pg_edu_pl/Documents/Microsoft%20Copilot%20Chat%20Files/cop_approach_gen.py)
POINTS_PER_MM = 2.83465
PAGE_W_PT = 595.275590551  # A4 width pt
PAGE_H_PT = 841.88976378   # A4 height pt

MARGIN_L_MM = 5
MARGIN_R_MM = 5
MARGIN_T_MM = 5
MARGIN_B_MM = 5

CONTENT_W_PT = (PAGE_W_PT - 2*MARGIN_L_MM*POINTS_PER_MM - 2*MARGIN_R_MM*POINTS_PER_MM)

COL_GAP_MM      = 6
TEXT_GAP        = 2  # (w generatorze używany do wysokości tytułów w punktach)
CHECKBOX_GAP_MM = 2
FONT_SIZE_LARGE = 12
FONT_SIZE_NORMAL= 8
FONT_SIZE_SMALL = 7

FIDUCIAL_SIZE_MM= 5
CHECKBOX_SIZE_MM= 5

CHECKBOX_SIZE_PT= CHECKBOX_SIZE_MM*POINTS_PER_MM
CHECKBOX_GAP_PT = CHECKBOX_GAP_MM*POINTS_PER_MM
COL_GAP_PT      = COL_GAP_MM*POINTS_PER_MM

CONTENT_PART_A_PT = CONTENT_W_PT/3.0
CONTENT_PART_B_PT = 2*CONTENT_W_PT/3.0 - COL_GAP_PT

X_A_PT = 2*MARGIN_L_MM*POINTS_PER_MM
Y_T_PT = PAGE_H_PT - MARGIN_T_MM*POINTS_PER_MM - (FONT_SIZE_LARGE + TEXT_GAP + 2*(FONT_SIZE_NORMAL + TEXT_GAP))
X_B_PT = 2*MARGIN_L_MM*POINTS_PER_MM + CONTENT_PART_A_PT + COL_GAP_PT

# Nagłówki skali w A: box_x = x0 + 12mm; box_dsc_y = Y_T - 28mm; y start = box_dsc_y - 2*FONT_SIZE_LARGE
BOX_X_OFFSET_MM  = 12
BOX_DSC_Y_OFFSET_MM = 28

ROW_LABEL_START_PT = (Y_T_PT - BOX_DSC_Y_OFFSET_MM*POINTS_PER_MM) - 2*FONT_SIZE_LARGE  # baseline etykiety kodu A.*.* [1](https://pgedupl-my.sharepoint.com/personal/micczube_pg_edu_pl/Documents/Microsoft%20Copilot%20Chat%20Files/cop_approach_gen.py)

# Wysokość/układ wiersza A (wynika z draw_part_a: przesunięcia po rysowaniu cyfr i kodu)
# Po każdej pozycji: y = by - CHECKBOX_SIZE; potem y -= CHECKBOX_GAP; w trakcie: y - 1.25*FONT_SIZE_SMALL (pod cyfrę)
ROW_STRIDE_PT = CHECKBOX_SIZE_PT + CHECKBOX_GAP_PT + 1.25*FONT_SIZE_SMALL  # ~5mm + 2mm + czcionka (w punktach) [1](https://pgedupl-my.sharepoint.com/personal/micczube_pg_edu_pl/Documents/Microsoft%20Copilot%20Chat%20Files/cop_approach_gen.py)

# Część B: y = Y_T - 24 (punkty); box_h = 45mm; QR: 20mm; linie co ~2*FONT_SIZE_LARGE
B_START_Y_PT = Y_T_PT - 24                                # baseline etykiety B.* (punkty) [1](https://pgedupl-my.sharepoint.com/personal/micczube_pg_edu_pl/Documents/Microsoft%20Copilot%20Chat%20Files/cop_approach_gen.py)
B_BOX_H_PT   = 45*POINTS_PER_MM
B_QR_SIZE_PT = 20*POINTS_PER_MM

@dataclass
class Box:
    x: float
    y: float  # bottom
    w: float
    h: float

# Fiduciale – środki (PDF-coords)
FIDUCIALS_TARGET_PDF = {
    'tl': (MARGIN_L_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2,
           PAGE_H_PT - (MARGIN_T_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2)),
    'tr': (PAGE_W_PT - (MARGIN_R_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2),
           PAGE_H_PT - (MARGIN_T_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2)),
    'br': (PAGE_W_PT - (MARGIN_R_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2),
           PAGE_H_PT - (MARGIN_B_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2)),
    'bl': (MARGIN_L_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2,
           PAGE_H_PT - (MARGIN_B_MM*POINTS_PER_MM + FIDUCIAL_SIZE_MM*POINTS_PER_MM/2)),
}

# ---------- Pozycje dla Części A ----------
# Generator rysuje sekcje w kolumnie A w stałym układzie pionowym; dla analizy przyjmujemy 20 kolejnych pozycji:
# Kolumny logicznie 7/7/6; w każdym wierszu kratki: zaczynają się od bx = x0 + 12mm, szerokość 6*box + 5*gap.
def build_positions_A() -> List[List[Box]]:
    positions_A: List[List[Box]] = []
    col_defs = [
        (X_A_PT + BOX_X_OFFSET_MM*POINTS_PER_MM,                      # bx bazowe (dla cyfr/kratek)
         X_A_PT,                                                      # x0 (lewy margines dla etykiety)
         CONTENT_PART_A_PT,                                           # szerokość kolumny A
         7),
        (X_A_PT + CONTENT_PART_A_PT + COL_GAP_PT + BOX_X_OFFSET_MM*POINTS_PER_MM,
         X_A_PT + CONTENT_PART_A_PT + COL_GAP_PT,
         CONTENT_PART_A_PT,
         7),
        (X_A_PT + 2*CONTENT_PART_A_PT + 2*COL_GAP_PT + BOX_X_OFFSET_MM*POINTS_PER_MM,
         X_A_PT + 2*CONTENT_PART_A_PT + 2*COL_GAP_PT,
         CONTENT_PART_A_PT,
         6),
    ]

    # baseline pierwszego wiersza w każdej kolumnie:
    start_label_y = ROW_LABEL_START_PT

    for bx_base, x0, col_w, n_rows in col_defs:
        for i in range(n_rows):
            label_y = start_label_y - i*ROW_STRIDE_PT
            # kratki dla skali 1..6: w generatorze by = label_y - 1mm; kratki = [bx_base + i*(size+gap), by-size]
            by = label_y - 1*POINTS_PER_MM
            boxes_row: List[Box] = []
            # szerokość 6-boxów:
            for j in range(6):
                x = bx_base + j*(CHECKBOX_SIZE_PT + CHECKBOX_GAP_PT)
                y = by - CHECKBOX_SIZE_PT
                boxes_row.append(Box(x=x, y=y, w=CHECKBOX_SIZE_PT, h=CHECKBOX_SIZE_PT))
            positions_A.append(boxes_row)

    return positions_A  # 20 pozycji

# ---------- Pozycje dla Części B ----------
def build_positions_B() -> List[Box]:
    positions_B: List[Box] = []
    x = X_B_PT
    y = B_START_Y_PT
    box_h = B_BOX_H_PT
    for _ in range(5):
        y_box_top = y - 2.5*POINTS_PER_MM
        rect = Box(x, y_box_top - box_h, CONTENT_PART_B_PT, box_h)
        positions_B.append(rect)
        # y przesuwany o box_h + 5mm (jak w generatorze)
        y = y_box_top - box_h - 5*POINTS_PER_MM
    return positions_B

# ---------- PDF -> image ----------
def pdf_first_page_to_image(path: str, dpi: int = 300) -> Image.Image:
    if convert_from_path is None:
        raise RuntimeError("Brak pdf2image. Zainstaluj: pip install pdf2image oraz Poppler.")
    imgs = convert_from_path(path, dpi=dpi)
    if not imgs:
        raise RuntimeError("Nie udało się przekonwertować PDF na obraz.")
    return imgs[0]  # tylko pierwsza strona

# ---------- Fiduciale: wyrównanie perspektywiczne ----------
def warp_to_template_perspective(pil_img: Image.Image, dpi: int = 300) -> Tuple[np.ndarray, int]:
    img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    h, w = img.shape[:2]
    pad = int(min(w, h) * 0.08)

    def region(corner):
        if corner == 'tl': return (0, h-pad, pad, h)
        if corner == 'tr': return (w-pad, h-pad, w, h)
        if corner == 'bl': return (0, 0, pad, pad)
        if corner == 'br': return (w-pad, 0, w, pad)

    def center_of_dark_square(r):
        x0, y0, x1, y1 = r
        roi = img[y0:y1, x0:x1]
        gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
        _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        contours, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            return (x0 + (x1-x0)//2, y0 + (y1-y0)//2)
        c = max(contours, key=cv2.contourArea)
        M = cv2.moments(c)
        if M['m00'] == 0:
            return (x0 + (x1-x0)//2, y0 + (y1-y0)//2)
        return (int(M['m10']/M['m00']) + x0, int(M['m01']/M['m00']) + y0)

    tl = center_of_dark_square(region('tl'))
    tr = center_of_dark_square(region('tr'))
    bl = center_of_dark_square(region('bl'))
    br = center_of_dark_square(region('br'))

    scale = dpi/72.0
    dst_pts_pdf = [
        FIDUCIALS_TARGET_PDF['tl'], FIDUCIALS_TARGET_PDF['tr'],
        FIDUCIALS_TARGET_PDF['br'], FIDUCIALS_TARGET_PDF['bl']
    ]
    dst_pts = np.array([(x*scale, y*scale) for (x, y) in dst_pts_pdf], dtype=np.float32)
    src_pts = np.array([tl, tr, br, bl], dtype=np.float32)

    out_w = int(PAGE_W_PT/72.0 * dpi)
    out_h = int(PAGE_H_PT/72.0 * dpi)
    M = cv2.getPerspectiveTransform(src_pts, dst_pts)
    warped = cv2.warpPerspective(img, M, (out_w, out_h), flags=cv2.INTER_CUBIC)
    return warped, dpi

# ---------- Detekcja kratki ----------
def read_checkbox(warped_bgr: np.ndarray, box_pt: Box, dpi: int,
                  thresh_method=cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU) -> float:
    scale = dpi/72.0
    x = int(box_pt.x * scale)
    y = int((PAGE_H_PT - box_pt.y - box_pt.h) * scale)  # invert Y
    w = int(box_pt.w * scale)
    h = int(box_pt.h * scale)
    roi = warped_bgr[y:y+h, x:x+w]
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(gray, 0, 255, thresh_method)
    inner = bw[2:-2, 2:-2] if bw.shape[0] > 4 and bw.shape[1] > 4 else bw
    dark_ratio = (inner > 0).sum() / inner.size
    return dark_ratio

# ---------- OCR pola B ----------
def ocr_field(warped_bgr: np.ndarray, rect_pt: Box, dpi: int) -> str:
    scale = dpi/72.0
    x = int(rect_pt.x * scale)
    y = int((PAGE_H_PT - rect_pt.y - rect_pt.h) * scale)
    w = int(rect_pt.w * scale)
    h = int(rect_pt.h * scale)
    roi = warped_bgr[y:y+h, x:x+w]
    gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
    _, thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    pil_img = Image.fromarray(thr)
    text = ""
    if pytesseract is not None:
        try:
            text = pytesseract.image_to_string(pil_img, lang='pol', config='--psm 6')
        except Exception:
            text = pytesseract.image_to_string(pil_img)
    return text.strip()

# ---------- Przetwarzanie 1 strony ----------
def process_first_page(pil_img: Image.Image):
    warped, dpi = warp_to_template_perspective(pil_img, dpi=300)
    positions_A = build_positions_A()
    positions_B = build_positions_B()

    answers_A: List[int] = []
    for boxes in positions_A:
        ratios = [read_checkbox(warped, b, dpi) for b in boxes]
        max_i = int(np.argmax(ratios))
        max_val = ratios[max_i]
        answers_A.append(max_i+1 if max_val > 0.15 else None)  # próg można dostroić

    answers_B_texts: List[str] = [ocr_field(warped, rect, dpi) for rect in positions_B]
    return answers_A, answers_B_texts

# ---------- Excel ----------
def save_to_excel(rows_a: List[List[int]], rows_b_text: List[List[str]], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wyniki'

    header = [f"A.{i}" for i in range(1,21)] + [
        'B.1', 'B.2', 'B.3', 'B.4', 'B.5'
    ]
    ws.append(header)

    for a_vals, b_texts in zip(rows_a, rows_b_text):
        row = [val if val is not None else '' for val in a_vals] + b_texts
        ws.append(row)

    from openpyxl.utils import get_column_letter
    for col in range(21, 26):  # 5 kolumn B
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 60
        for row_idx in range(2, ws.max_row+1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    # Wiersze podsumowań: średnia/odchylenie bez '6' oraz liczba '6'
    def numeric_column_values(col_idx):
        vals = []
        for r in range(2, ws.max_row+1):
            v = ws.cell(row=r, column=col_idx).value
            if isinstance(v, (int, float)):
                vals.append(v)
            elif isinstance(v, str) and v.strip().isdigit():
                vals.append(int(v.strip()))
        return vals

    mean_row   = ['']*25
    std_row    = ['']*25
    count6_row = ['']*25
    for col in range(1, 21):
        vals = numeric_column_values(col)
        excl = [x for x in vals if x != 6]
        if excl:
            mean_row[col-1] = float(np.mean(excl))
            std_row[col-1]  = float(np.std(excl, ddof=1)) if len(excl) > 1 else 0.0
        count6_row[col-1] = int(sum(1 for x in vals if x == 6))

    ws.append(mean_row)
    ws.append(std_row)
    ws.append(count6_row)
    ws.cell(row=ws.max_row-2, column=1, value='Średnia (bez 6)')
    ws.cell(row=ws.max_row-1, column=1, value='Odchylenie standardowe (bez 6)')
    ws.cell(row=ws.max_row,   column=1, value="Liczba odpowiedzi '6'")

    wb.save(output_path)

# ---------- Iteracja po wejściu: folder lub jeden PDF; tylko pierwsze strony ----------
def iter_first_pages(input_path: str, dpi: int = 300):
    if os.path.isdir(input_path):
        for name in sorted(os.listdir(input_path)):
            if name.lower().endswith('.pdf'):
                full = os.path.join(input_path, name)
                try:
                    img = pdf_first_page_to_image(full, dpi=dpi)
                    yield (name, img)
                except Exception as e:
                    print(f"Błąd konwersji {name}: {e}", file=sys.stderr)
    else:
        if not input_path.lower().endswith('.pdf'):
            raise ValueError('Wejście musi być PDF-em lub folderem z PDF-ami.')
        img = pdf_first_page_to_image(input_path, dpi=dpi)
        yield (os.path.basename(input_path) + "#1", img)

def main(input_file='./20260108130546954.pdf', output_file='./wyniki.xlsx'):
    # parser = argparse.ArgumentParser(description='Analiza pierwszych stron ankiet (student/teacher).')
    # parser.add_argument('--input', required=True, help='Folder z PDF-ami lub pojedynczy PDF (wiele stron).')
    # parser.add_argument('--output', default='wyniki.xlsx', help='Plik wyjściowy Excel.')
    # args = parser.parse_args()

    rows_a: List[List[int]] = []
    rows_b_text: List[List[str]] = []
    pages = 0

    for name, pil_img in iter_first_pages(input_file, dpi=300):
        try:
            a_vals, b_texts = process_first_page(pil_img)
            rows_a.append(a_vals)
            rows_b_text.append(b_texts)
            pages += 1
            print(f"OK: {name}")
        except Exception as e:
            print(f"Błąd strony {name}: {e}", file=sys.stderr)

    if pages == 0:
        print('Brak danych do zapisu.', file=sys.stderr)
        sys.exit(2)

    save_to_excel(rows_a, rows_b_text, output_file)
    print(f"Zapisano wyniki do: {output_file}")

if __name__ == '__main__':
    main(input_file='./20260108130546954.pdf', output_file='./wyniki.xlsx')
