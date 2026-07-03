import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import io

part_a_texts = {
    'A.1.1': 'Zajęcia były prowadzone w sposób zrozumiały i uporządkowany.',
    'A.1.2': 'Tempo zajęć nie było odpowiednie dla poziomu grupy.',
    'A.1.3': 'Prowadzący zachęcał do zadawania pytań i aktywnego udziału.',
    'A.1.4': 'Prowadzący nie był przygotowany merytorycznie do prowadzenia zajęć.',
    'A.1.5': 'Prowadzący dostosowywał treści do uczestników zajęć.',
    'A.2.1': 'Materiały udostępniane do zajęć były przydatne.',
    'A.2.2': 'Materiały udostępniane do zajęć były przestarzałe.',
    'A.2.3': 'Zajęcia nie były zgodne z kartą przedmiotu i zapowiedzianymi treściami.',
    'A.2.4': 'Forma zajęć (W/Ć/L/P/S) sprzyjała przyswajaniu wiedzy.',
    'A.2.5': 'Wymagania dotyczące zaliczenia nie były jasno określone.',
    'A.3.1': 'Prowadzący przekazywał informacje w sposób przystępny.',
    'A.3.2': 'Prowadzący był zamknięty na opinie i sugestie studentów',
    'A.3.3': 'Atmosfera na zajęciach sprzyjała uczeniu się.',
    'A.3.4': 'Komunikacja między prowadzącym a studentami była jasna i kulturalna.',
    'A.4.1': 'Treści realizowane na zajęciach były interesujące.',
    'A.4.2': 'Treści zajęć były powiązane z praktyką inżynierską.',
    'A.4.3': 'Stopień zaawansowania grupy był adekwatny do prowadzonych zajęć.',
    'A.4.4': 'Poziom trudności zajęć był adekwatny do mojego przygotowania.',
    'A.4.5': 'Zajęcia nie przyczyniły się do zwiększenia moich kompetencji.',
    'A.4.6': 'Oceniam ten przedmiot jako wartościowy dla mojego kierunku studiów.',
}

part_b_texts = [
    'B.1. Co było najmocniejszą stroną tych zajęć?',
    'B.2. Co było najsłabszą stroną tych zajęć?',
    'B.3. Co należałoby poprawić w sposobie prowadzenia zajęć lub organizacji przedmiotu?',
    'B.4. Jakie elementy zajęć były dla Ciebie najbardziej przydatne lub inspirujące?',
    'B.5. Jakie elementy merytoryczne należałoby dodać do przedmiotu?',
]

# Part A: int keys 11-15, 21-25, 31-34, 41-46 -> column index (0-based)
_A_OFFSETS = {1: 0, 2: 5, 3: 10, 4: 14}
def a_key_to_col_idx(k: int) -> int:
    group, item = divmod(k, 10)
    return _A_OFFSETS[group] + (item - 1)

# Part B: string keys 'b1'-'b5' -> index 0-4
def b_key_to_idx(k: str) -> int:
    return int(k[1]) - 1

def numpy_to_image_bytes(arr: np.ndarray) -> io.BytesIO:
    if arr.dtype != np.uint8:
        arr = (arr * 255).clip(0, 255).astype(np.uint8)
    buf = io.BytesIO()
    Image.fromarray(arr).save(buf, format='PNG')
    buf.seek(0)
    return buf


def build_excel(respondents: list, output_path: str = 'report.xlsx'):
    """
    respondents: list of merged dicts, one per respondent.
      Part A keys: int  11-15, 21-25, 31-34, 41-46  -> int value 1-6
      Part B keys: str  'b1'-'b5'                    -> tuple (str, np.ndarray)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raport'

    yellow = PatternFill('solid', start_color='FFFF00')
    bold9  = Font(name='Arial', bold=True, size=9)
    reg9   = Font(name='Arial', size=9)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    rotated  = Alignment(horizontal='center', vertical='bottom', text_rotation=90, wrap_text=True)
    left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin   = Side(style='thin')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_a, n_b = len(part_a_texts), len(part_b_texts)   # 20, 5
    COL_LABEL   = 1
    COL_A       = 2              # cols 2..21
    COL_B_TXT   = COL_A + n_a   # cols 22..26
    COL_B_IMG   = COL_B_TXT + n_b  # cols 27..31

    # ── Row 1: headers ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 120
    ws.cell(1, COL_LABEL, '#').font = bold9

    for i, (key, text) in enumerate(part_a_texts.items()):
        c = ws.cell(1, COL_A + i, f"{key}: {text}")
        c.font, c.alignment, c.border = bold9, rotated, brd

    for j, txt in enumerate(part_b_texts):
        c = ws.cell(1, COL_B_TXT + j, txt)
        c.font, c.alignment, c.border = bold9, rotated, brd
        c = ws.cell(1, COL_B_IMG + j, f"[wykres] {txt}")
        c.font, c.alignment, c.border = bold9, rotated, brd

    # ── Row 2: averages (yellow) ─────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    c = ws.cell(2, COL_LABEL, 'Średnia')
    c.font, c.alignment = bold9, center

    n = len(respondents)
    r0, r1 = 3, 2 + n  # data row range

    for i in range(n_a):
        col = COL_A + i
        ltr = get_column_letter(col)
        c = ws.cell(2, col, f'=IFERROR(AVERAGE({ltr}{r0}:{ltr}{r1}),"")')
        c.font, c.fill, c.alignment, c.border = bold9, yellow, center, brd
        c.number_format = '0.00'

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, respondent in enumerate(respondents):
        xrow = 3 + row_idx
        ws.row_dimensions[xrow].height = 80

        c = ws.cell(xrow, COL_LABEL, row_idx + 1)
        c.font, c.alignment = reg9, center

        for key, value in respondent.items():
            if isinstance(key, int):                      # Part A
                col = COL_A + a_key_to_col_idx(key)
                c = ws.cell(xrow, col, 'N/A' if value == 6 else value)
                c.font, c.alignment, c.border = reg9, center, brd

            elif isinstance(key, str) and key.startswith('b'):  # Part B
                j = b_key_to_idx(key)
                text, img_arr = value

                c = ws.cell(xrow, COL_B_TXT + j, text)
                c.font, c.alignment, c.border = reg9, left_top, brd

                img = openpyxl.drawing.image.Image(numpy_to_image_bytes(img_arr))
                img.anchor = f"{get_column_letter(COL_B_IMG + j)}{xrow}"
                img.width, img.height = 150, 70
                ws.add_image(img)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_LABEL)].width = 5
    for i in range(n_a):
        ws.column_dimensions[get_column_letter(COL_A + i)].width = 6
    for j in range(n_b):
        ws.column_dimensions[get_column_letter(COL_B_TXT + j)].width = 30
        ws.column_dimensions[get_column_letter(COL_B_IMG + j)].width = 22

    wb.save(output_path)
    print(f"Saved: {output_path}")


# ── Demo ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    rng = np.random.default_rng(42)

    a_keys = list(range(11, 16)) + list(range(21, 26)) + list(range(31, 35)) + list(range(41, 47))
    b_keys = [f'b{i}' for i in range(1, 6)]

    def fake_img(seed):
        r = np.random.default_rng(seed)
        return r.integers(0, 256, (80, 150, 3), dtype=np.uint8)

    respondents = []
    for i in range(10):
        d = {k: int(rng.integers(1, 7)) for k in a_keys}
        d.update({k: (f"Odpowiedź respondenta {i+1} na pytanie {k.upper()}.", fake_img(i * 10 + int(k[1])))
                  for k in b_keys})
        respondents.append(d)

    build_excel(respondents, './report.xlsx')