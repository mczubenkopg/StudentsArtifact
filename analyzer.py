"""
Analizator wypełnionych ankiet studenckich
Wymagania: pip install opencv-python pytesseract pandas Pillow pdf2image openpyxl numpy
"""

from pathlib import Path

import cv2
import numpy as np
import pytesseract
from pdf2image import convert_from_path
import pandas as pd
from pyzbar.pyzbar import decode
from pathlib import Path
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import re

from qreader import QReader


class SurveyAnalyzer:
    def __init__(self, input_folder, output_excel):
        self.input_folder = Path(input_folder)
        self.output_excel = output_excel
        self.results = []

        # Definicja kolumn zgodnie z kodem generatora
        self.part_a_columns = [
            'A.1.1', 'A.1.2', 'A.1.3', 'A.1.4', 'A.1.5',
            'A.2.1', 'A.2.2', 'A.2.3', 'A.2.4', 'A.2.5',
            'A.3.1', 'A.3.2', 'A.3.3', 'A.3.4',
            'A.4.1', 'A.4.2', 'A.4.3', 'A.4.4', 'A.4.5', 'A.4.6'
        ]

        self.part_b_columns = ['B.1', 'B.2', 'B.3', 'B.4', 'B.5']

        # Pytania dla nagłówków
        self.questions_text = {
            'A.1.1': 'Zajęcia były prowadzone w sposób zrozumiały i uporządkowany.',
            'A.1.2': 'Tempo zajęć nie było odpowiednie dla poziomu grupy.',
            'A.1.3': 'Prowadzący zachęcał do zadawania pytań i aktywnego udziału.',
            'A.1.4': 'Prowadzący nie był przygotowany merytorycznie do prowadzenia zajęć.',
            'A.1.5': 'Prowadzący dostosowywał treści do uczestników zajęć.',
            'A.2.1': 'Materiały udostępniane do zajęć były przydatne.',
            'A.2.2': 'Materiały udostępniane do zajęć były przestarzałe.',
            'A.2.3': 'Zajęcia nie były zgodne z kartą przedmiotu i zapowiedzianymi treściami.',
            'A.2.4': 'Forma zajęć (W/Ć/L/P/S) sprzyjała przyswajaniu wiedzy.',
            'A.2.5': 'Wymagania dotyczące zaliczenia nie były jasno określone.',
            'A.3.1': 'Prowadzący przekazywał informacje w sposób przystępny.',
            'A.3.2': 'Prowadzący był zamknięty na opinie i sugestie studentów',
            'A.3.3': 'Atmosfera na zajęciach sprzyjała uczeniu się.',
            'A.3.4': 'Komunikacja między prowadzącym a studentami była jasna i kulturalna.',
            'A.4.1': 'Treści realizowane na zajęciach były interesujące.',
            'A.4.2': 'Treści zajęć były powiązane z praktyką inżynierską.',
            'A.4.3': 'Stopień zaawansowania grupy był adekwatny do prowadzonych zajęć.',
            'A.4.4': 'Poziom trudności zajęć był adekwatny do mojego przygotowania.',
            'A.4.5': 'Zajęcia nie przyczyniły się do zwiększenia moich kompetencji.',
            'A.4.6': 'Oceniam ten przedmiot jako wartościowy dla mojego kierunku studiów.',
            'B.1': 'Co było najmocniejszą stroną tych zajęć?',
            'B.2': 'Co było najsłabszą stroną tych zajęć?',
            'B.3': 'Co należałoby poprawić w sposobie prowadzenia zajęć lub organizacji przedmiotu?',
            'B.4': 'Jakie elementy zajęć były dla Ciebie najbardziej przydatne lub inspirujące?',
            'B.5': 'Jakie elementy merytoryczne należałoby dodać do przedmiotu?'
        }

    def group_checkboxes_by_row(self, checkboxes):
        # simple heuristic: group by Y coordinate
        checkboxes.sort(key=lambda x: x["bbox"][1])

        rows = []
        current_row = []

        threshold = 10  # px

        for cb in checkboxes:
            if not current_row:
                current_row.append(cb)
                continue

            if abs(cb["bbox"][1] - current_row[-1]["bbox"][1]) < threshold:
                current_row.append(cb)
            else:
                rows.append(current_row)
                current_row = [cb]

        if current_row:
            rows.append(current_row)

        return rows

    def analyze_surveys(self):
        """Analizuje wszystkie pliki PDF w folderze"""

        pdf_files = list(self.input_folder.glob("*.pdf"))

        print(f"Znaleziono {len(pdf_files)} plików PDF do analizy...")

        for pdf_file in pdf_files:
            print(f"\nPrzetwarzanie: {pdf_file.name}")
            try:
                # Konwersja PDF na obrazy (może być wielostronicowy)
                images = convert_from_path(pdf_file, dpi=300)
                print(f"  Liczba stron: {len(images)}")
                for page_num, img in enumerate(images[1:]):
                    print(f"  Analiza strony {page_num}...")
                    try:
                        result = self.process_single_page(img)
                        if result:
                            result['plik'] = pdf_file.name
                            result['strona'] = page_num
                            self.results.append(result)
                            print(f"    ✓ Strona {page_num} przeanalizowana")
                    except Exception as e:
                        print(f"    ✗ Błąd strony {page_num}: {str(e)}")

            except Exception as e:
                print(f"✗ Błąd przy przetwarzaniu {pdf_file.name}: {str(e)}")

        if self.results:
            self.save_to_excel()
            print(f"\n✓ Wyniki zapisano do {self.output_excel}")
            print(f"  Przeanalizowano {len(self.results)} stron ankiet")
        else:
            print("\n✗ Brak wyników do zapisania")

    def process_single_page(self, img):
        """Przetwarza pojedynczą stronę ankiety"""
        result = {}
        # Analiza części A - kratki do zakreślenia
        closed_answers = self.analyze_closed_questions(img)
        result.update(closed_answers)

        # # Analiza części B - pytania otwarte
        # open_answers = self.analyze_open_questions(gray)
        # result.update(open_answers)

        return result

    # def find_fiducials(self, gray_img):
    #     """Znajduje znaczniki fiducjalne w rogach strony"""
    #     height, width = gray_img.shape
    #     _, binary = cv2.threshold(gray_img, 127, 255, cv2.THRESH_BINARY_INV)
    #
    #     # Szukamy czarnych kwadratów w rogach
    #     margin = int(height * 0.05)
    #     fiducial_size = int(height * 0.02)
    #
    #     corners = {
    #         'top_left': (0, 0, margin, margin),
    #         'top_right': (width - margin, 0, width, margin),
    #         'bottom_left': (0, height - margin, margin, height),
    #         'bottom_right': (width - margin, height - margin, width, height)
    #     }
    #     fiducials = {}
    #     for corner_name, (x1, y1, x2, y2) in corners.items():
    #         roi = binary[y1:y2, x1:x2]
    #         contours, _ = cv2.findContours(roi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    #
    #         for contour in contours:
    #             area = cv2.contourArea(contour)
    #             if area > 100:
    #                 x, y, w, h = cv2.boundingRect(contour)
    #                 if 0.8 < w/h < 1.2:  # Prawie kwadrat
    #                     fiducials[corner_name] = (x1 + x + w//2, y1 + y + h//2)
    #                     break
    #
    #     return fiducials

    def analyze_closed_questions(self, img):
        """Analizuje zakreślone kratki w pytaniach zamkniętych (1-6)"""
        answers = {}
        qreader = QReader(model_size='s', min_confidence=0.5)
        # Get the image that contains the QR code
        left_img = np.array(img)[:,:int(img.size[0]/3),:]
        image = cv2.cvtColor(np.array(img)[:,:int(img.size[0]/3),:], cv2.COLOR_BGR2RGB)

        draw = ImageDraw.Draw(img)
        # Use the detect_and_decode function to get the decoded QR data
        b_codes = qreader.detect(image=image)
        # to łapie wszystkie bcody
        print(b_codes)
        pyzbar_out = decode(image=left_img)
        #to działa w kontekście numerowanych kodów (nie wszystkich!!)
        print(pyzbar_out)
        # pyzbar_out = tuple(out.data.data.decode('utf-8').encode('shift-jis').decode('utf-8') for out in pyzbar_out)

        for b_code in b_codes:
            draw.rectangle(b_code.get('bbox_xyxy'), outline='red', fill='green', width=3)
            # image = cv2.rectangle(image, (x0, y0), (x1, y1), (0, 255, 0), 2)
        img.show('Image')
        pass

    # Znajdź znaczniki do kalibracji
    #     fiducials = self.find_fiducials(gray_img)
    #
    #     # Binaryzacja
    #     _, binary = cv2.threshold(gray_img, 127, 255, cv2.THRESH_BINARY_INV)
    #
    #     # Parametry layoutu z generatora (przeliczone na piksele przy 300 DPI)
    #     dpi = 300
    #     mm_to_px = dpi / 25.4
    #
    #     # Wymiary z generatora
    #     margin_l = 5 * mm_to_px
    #     content_w = (210 - 10) * mm_to_px  # A4 width minus margins
    #     content_part_a = content_w / 3
    #
    #     checkbox_size = 5 * mm_to_px
    #     checkbox_gap = 2 * mm_to_px
    #
    #     # Lokalizacja kratek - prawa strona części A
    #     x_start = margin_l + 12 * mm_to_px
    #     y_start = height * 0.15  # Przybliżona lokalizacja początkowa
    #
    #     # Wykrywanie wszystkich małych kwadratów (kratek)
    #     contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    #
    #     checkboxes = []
    #     for contour in contours:
    #         x, y, w, h = cv2.boundingRect(contour)
    #
    #         # Filtrowanie: małe kwadraty w odpowiednim obszarze
    #         if (checkbox_size * 0.7 < w < checkbox_size * 1.3 and
    #             checkbox_size * 0.7 < h < checkbox_size * 1.3 and
    #             x > width * 0.15 and x < width * 0.45 and
    #             y > height * 0.12):
    #
    #             area = cv2.contourArea(contour)
    #             if area > 50:
    #                 checkboxes.append((x, y, w, h))
    #
    #     # Sortowanie kratek: najpierw po Y (rzędy), potem po X (kolumny)
    #     checkboxes.sort(key=lambda b: (b[1], b[0]))
    #
    #     # Grupowanie kratek w rzędy (po 6 kratek w rzędzie)
    #     rows = []
    #     current_row = []
    #     last_y = -100
    #
    #     for box in checkboxes:
    #         x, y, w, h = box
    #
    #         if abs(y - last_y) < checkbox_size * 0.5:
    #             current_row.append(box)
    #         else:
    #             if len(current_row) >= 4:  # Minimum 4 kratki w rzędzie
    #                 rows.append(sorted(current_row, key=lambda b: b[0])[:6])
    #             current_row = [box]
    #         last_y = y
    #
    #     if len(current_row) >= 4:
    #         rows.append(sorted(current_row, key=lambda b: b[0])[:6])
    #
    #     # Przypisanie do pytań (20 pytań w części A)
    #     for idx, row in enumerate(rows[:20]):
    #         if idx >= len(self.part_a_columns):
    #             break
    #
    #         q_id = self.part_a_columns[idx]
    #
    #         # Sprawdzenie wypełnienia każdej kratki
    #         max_fill = 0
    #         selected = None
    #
    #         for box_idx, (x, y, w, h) in enumerate(row):
    #             # Region kratki
    #             roi = binary[y:y+h, x:x+w]
    #             if roi.size == 0:
    #                 continue
    #
    #             # Procent wypełnienia
    #             fill_ratio = np.sum(roi > 0) / roi.size
    #
    #             if fill_ratio > max_fill:
    #                 max_fill = fill_ratio
    #                 selected = box_idx + 1
    #
    #         # Próg wypełnienia - 30%
    #         if max_fill > 0.3:
    #             answers[q_id] = selected
    #         else:
    #             answers[q_id] = None

        return answers

    # def analyze_open_questions(self, gray_img):
    #     """Analizuje pytania otwarte z OCR"""
    #     results = {}
    #     height, width = gray_img.shape
    #
    #     # Binaryzacja dla lepszego OCR
    #     _, binary = cv2.threshold(gray_img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    #
    #     # Szukanie dużych prostokątów (ramki dla odpowiedzi)
    #     contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    #
    #     rectangles = []
    #     for contour in contours:
    #         x, y, w, h = cv2.boundingRect(contour)
    #
    #         # Duże prostokąty po prawej stronie
    #         if (w > width * 0.45 and
    #             100 < h < int(height * 0.25) and
    #             x > width * 0.4):
    #             rectangles.append((x, y, w, h, y))
    #
    #     rectangles.sort(key=lambda r: r[4])
    #
    #     # Przetwarzanie pierwszych 5 ramek
    #     for idx, (x, y, w, h, _) in enumerate(rectangles[:5]):
    #         if idx >= len(self.part_b_columns):
    #             break
    #
    #         q_id = self.part_b_columns[idx]
    #
    #         # Wycinanie regionu
    #         padding = 10
    #         roi = gray_img[max(0, y+padding):min(height, y+h-padding),
    #                       max(0, x+padding):min(width, x+w-padding)]
    #
    #         # Preprocessing dla lepszego OCR
    #         roi = cv2.bilateralFilter(roi, 9, 75, 75)
    #         _, roi_binary = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    #
    #         # OCR z polskim językiem
    #         config = '--oem 3 --psm 6 -l pol'
    #         text = pytesseract.image_to_string(roi_binary, config=config)
    #
    #         # Czyszczenie tekstu
    #         text = text.strip()
    #         text = re.sub(r'\s+', ' ', text)
    #
    #         results[q_id] = text if text else ''
    #
    #     # Uzupełnienie brakujących pytań
    #     for q_id in self.part_b_columns:
    #         if q_id not in results:
    #             results[q_id] = ''
    #
    #     return results

    # def save_to_excel(self):
    #     """Zapisuje wyniki do pliku Excel z formatowaniem"""
    #     # Tworzenie DataFrame
    #     data_rows = []
    #
    #     for result in self.results:
    #         row = {
    #             'Plik': result.get('plik', ''),
    #             'Strona': result.get('strona', '')
    #         }
    #
    #         # Pytania zamknięte (część A)
    #         for q in self.part_a_columns:
    #             row[q] = result.get(q, None)
    #
    #         # Pytania otwarte (część B)
    #         for q in self.part_b_columns:
    #             row[q] = result.get(q, '')
    #
    #         data_rows.append(row)
    #
    #     df = pd.DataFrame(data_rows)
    #
    #     # Zapisanie do Excel
    #     with pd.ExcelWriter(self.output_excel, engine='openpyxl') as writer:
    #         df.to_excel(writer, sheet_name='Wyniki', index=False)
    #
    #         workbook = writer.book
    #         worksheet = writer.sheets['Wyniki']
    #
    #         # Formatowanie
    #         self.format_excel(worksheet, df)
    #
    # def format_excel(self, worksheet, df):
    #     """Formatuje arkusz Excel"""
    #     # Kolory nagłówków
    #     header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    #     header_font = Font(bold=True, color='FFFFFF')
    #
    #     # Formatowanie nagłówka
    #     for cell in worksheet[1]:
    #         cell.fill = header_fill
    #         cell.font = header_font
    #         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    #
    #     # Dodanie wiersza z pełnymi pytaniami
    #     worksheet.insert_rows(2)
    #     worksheet['A2'] = 'Treść pytania'
    #     worksheet['B2'] = ''
    #
    #     col_idx = 3
    #     for q_id in self.part_a_columns + self.part_b_columns:
    #         cell = worksheet.cell(row=2, column=col_idx)
    #         cell.value = self.questions_text.get(q_id, '')
    #         cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    #         col_idx += 1
    #
    #     # Szerokości kolumn
    #     worksheet.column_dimensions['A'].width = 20
    #     worksheet.column_dimensions['B'].width = 8
    #
    #     # Kolumny część A (wąskie)
    #     for col_idx in range(3, 3 + len(self.part_a_columns)):
    #         worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 12
    #
    #     # Kolumny część B (szerokie dla tekstu)
    #     for col_idx in range(3 + len(self.part_a_columns), 3 + len(self.part_a_columns) + len(self.part_b_columns)):
    #         col_letter = worksheet.cell(row=1, column=col_idx).column_letter
    #         worksheet.column_dimensions[col_letter].width = 50
    #
    #     # Formatowanie komórek z pytaniami otwartymi
    #     for row in range(3, worksheet.max_row + 1):
    #         for col_idx in range(3 + len(self.part_a_columns), 3 + len(self.part_a_columns) + len(self.part_b_columns)):
    #             cell = worksheet.cell(row=row, column=col_idx)
    #             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    #
    #     # Wysokość wiersza z pytaniami
    #     worksheet.row_dimensions[2].height = 60
    #
    #     # Wysokość wierszy z odpowiedziami
    #     for row in range(3, worksheet.max_row + 1):
    #         worksheet.row_dimensions[row].height = 80
    #
    #     # Dodanie statystyk na końcu
    #     last_row = worksheet.max_row + 2
    #
    #     # Średnia (bez wartości 6)
    #     worksheet.cell(row=last_row, column=1, value='Średnia (bez 6)')
    #     worksheet.cell(row=last_row, column=1).font = Font(bold=True)
    #
    #     for col_idx, q_id in enumerate(self.part_a_columns, start=3):
    #         col_letter = worksheet.cell(row=1, column=col_idx).column_letter
    #         # Formuła: średnia z zakresu, ignorując wartość 6 i puste
    #         formula = f'=AVERAGEIFS({col_letter}3:{col_letter}{worksheet.max_row-2},{col_letter}3:{col_letter}{worksheet.max_row-2},"<>6",{col_letter}3:{col_letter}{worksheet.max_row-2},"<>")'
    #         worksheet.cell(row=last_row, column=col_idx, value=formula)
    #         worksheet.cell(row=last_row, column=col_idx).number_format = '0.00'
    #
    #     last_row += 1
    #
    #     # Odchylenie standardowe (bez wartości 6)
    #     worksheet.cell(row=last_row, column=1, value='Odchylenie std. (bez 6)')
    #     worksheet.cell(row=last_row, column=1).font = Font(bold=True)
    #
    #     for col_idx, q_id in enumerate(self.part_a_columns, start=3):
    #         col_letter = worksheet.cell(row=1, column=col_idx).column_letter
    #         # Ręczne obliczenie odchylenia dla wartości różnych od 6
    #         # Używamy pomocniczych formuł
    #         cell = worksheet.cell(row=last_row, column=col_idx)
    #         # Dla uproszczenia - używamy STDEV.S na odfiltrowanych danych
    #         # W Excel trudno to zrobić jedną formułą, więc robimy aproximację
    #         formula = f'=STDEV.S(IF(({col_letter}3:{col_letter}{worksheet.max_row-3}<>6)*({col_letter}3:{col_letter}{worksheet.max_row-3}<>""),{col_letter}3:{col_letter}{worksheet.max_row-3}))'
    #         # To wymaga formuły tablicowej - w openpyxl zapisujemy jako tekst
    #         cell.value = f'=STDEV({col_letter}3:{col_letter}{worksheet.max_row-3})'
    #         cell.number_format = '0.00'
    #
    #     last_row += 1
    #
    #     # Liczba odpowiedzi "6"
    #     worksheet.cell(row=last_row, column=1, value='Liczba odpowiedzi "6"')
    #     worksheet.cell(row=last_row, column=1).font = Font(bold=True)
    #
    #     for col_idx, q_id in enumerate(self.part_a_columns, start=3):
    #         col_letter = worksheet.cell(row=1, column=col_idx).column_letter
    #         formula = f'=COUNTIF({col_letter}3:{col_letter}{worksheet.max_row-3},6)'
    #         worksheet.cell(row=last_row, column=col_idx, value=formula)
    #
    #     # Zamrożenie górnych wierszy
    #     worksheet.freeze_panes = 'A3'


def main():
    """Główna funkcja programu"""
    # Konfiguracja
    INPUT_FOLDER = "./sample"
    OUTPUT_EXCEL = "wyniki_ankiet.xlsx"
    print("=" * 70)
    print("ANALIZATOR ANKIET STUDENCKICH")
    print("=" * 70)
    print("\nKonfiguracja:")
    print(f"  Folder wejściowy: {INPUT_FOLDER}/")
    print(f"  Plik wyjściowy:   {OUTPUT_EXCEL}")
    print(f"  Format:           21 kolumn + 5 kolumn tekstowych")
    print(f"                    20 pytań części A + 5 pytań części B")

    print("\nWymagania:")
    print("  - Skany w folderze jako pliki PDF (jedno- lub wielostronicowe)")
    print("  - Tesseract OCR zainstalowany w systemie")
    print("  - Pakiety: opencv-python, pytesseract, pandas, pdf2image, openpyxl")
    print("=" * 70)

    # Uruchomienie analizy
    analyzer = SurveyAnalyzer(INPUT_FOLDER, OUTPUT_EXCEL)
    analyzer.analyze_surveys()

    print("\n" + "=" * 70)
    print("ANALIZA ZAKOŃCZONA")
    print("=" * 70)


if __name__ == "__main__":
    main()